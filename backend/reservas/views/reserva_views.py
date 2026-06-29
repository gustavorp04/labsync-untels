from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from django.db import transaction
from django.core.cache import cache
from datetime import timedelta
from ..serializers.reserva_serializers import CrearReservaSerializer, HorarioDisponibleSerializer
from ..services import reserva_service
from ..models import Usuario, Penalizacion, HorarioDisponible, Laboratorio, ActivoLaboratorio, Reserva, ReservaDetalle, HistorialReserva, Asistencia
import logging
import csv
from django.http import HttpResponse
from django.db.models import Count, Q
from ..utils.auth import IsAdminOrJefatura, IsJefatura  # ID-05/06

logger = logging.getLogger('reservas')

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def horarios_por_laboratorio(request, id_laboratorio):
    """Devuelve los horarios de un laboratorio específico.
    Si se proporciona el parámetro 'fecha' en la consulta, se filtrará exactamente por esa fecha
    (mostrando todos los horarios, útil para la vista del administrador).
    De lo contrario, se devuelven los horarios futuros que estén 'Disponible' (para reserva).
    """
    fecha = request.query_params.get('fecha')
    if fecha:
        # N-4: select_related evita N+1 al serializar tipo_nombre y aforo_maximo
        horarios = (
            HorarioDisponible.objects
            .filter(id_laboratorio=id_laboratorio, fecha=fecha)
            .select_related('id_laboratorio', 'id_laboratorio__id_tipo')
            .order_by('hora_inicio')
        )
    else:
        hoy = timezone.localtime(timezone.now()).date()
        # N-4: select_related evita N+1 al serializar tipo_nombre y aforo_maximo
        horarios = (
            HorarioDisponible.objects
            .filter(
                id_laboratorio=id_laboratorio,
                fecha__gte=hoy,
                estado='Disponible',
            )
            .select_related('id_laboratorio', 'id_laboratorio__id_tipo')
            .order_by('fecha', 'hora_inicio')
        )
    
    serializer = HorarioDisponibleSerializer(horarios, many=True)
    return Response(serializer.data)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def mis_reservas(request, id_usuario):
    """
    GET: Devuelve las reservas del usuario indicado por id_usuario en la ruta anidada.
    POST: Crea una reserva para el usuario indicado por id_usuario en la ruta anidada
          (estudiante o docente según el rol del usuario).
    """
    # Validar permisos: el usuario autenticado debe coincidir con el id_usuario de la ruta (o ser admin/jefatura)
    if str(request.user.id_usuario) != str(id_usuario) and request.user.id_rol.nombre not in ('admin_lab', 'jefatura'):
        return Response({'error': 'No tienes permiso para acceder a las reservas de este usuario.'}, status=403)

    try:
        usuario = Usuario.objects.get(pk=id_usuario)
    except Usuario.DoesNotExist:
        return Response({'error': "Usuario no encontrado."}, status=404)

    if request.method == 'GET':
        from ..serializers.reserva_serializers import ReservaSerializer
        # A-1: select_related + prefetch_related eliminan el N+1 del serializer
        reservas = (
            Reserva.objects
            .filter(id_usuario=id_usuario)
            .select_related(
                'id_usuario', 'id_usuario__id_rol',
                'id_horario', 'id_horario__id_laboratorio',
            )
            .prefetch_related(
                'reservadetalle_set__id_activo__id_tipo_activo',
                'historialreserva_set',
            )
            .order_by('-created_at')
        )
        serializer = ReservaSerializer(reservas, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        # S-7 (CWE-639 / IDOR): la reserva se crea SIEMPRE para el usuario
        # autenticado (request.user), nunca para el id_usuario de la URL ni para
        # un user_id del payload — ambos los controla el cliente. El guard de
        # arriba ya bloquea el acceso cruzado; vincular aquí a request.user es
        # defensa en profundidad: aunque ese guard se relajara (p. ej. para que
        # un admin liste reservas ajenas), la creación nunca podrá suplantar a
        # otro usuario. El rol que decide el flujo también es el del autenticado.
        usuario = request.user
        rol = usuario.id_rol.nombre
        if rol == 'docente':
            serializer = CrearReservaSerializer(data=request.data)
            if not serializer.is_valid():
                primer_error = list(serializer.errors.values())[0][0]
                return Response({'error': str(primer_error)}, status=400)

            data = serializer.validated_data
            reserva, error = reserva_service.crear_reserva_docente(
                usuario=usuario,
                horario=data['horario_obj'],
                cantidad_alumnos=data['cantidad_alumnos'],
                acepta_dj=data['acepto_declaracion_jurada'],
                activos_ids=request.data.get('activos_ids', [])
            )
            if error:
                return Response({'error': error}, status=409)

            horario_obj = data['horario_obj']
            cache.delete(f"activos_lab_{horario_obj.id_laboratorio_id}_{horario_obj.id_horario}")

            return Response({
                'mensaje': 'Reserva creada con éxito.',
                'reserva_id': reserva.id_reserva
            }, status=201)

        elif rol == 'estudiante':
            try:
                reserva_service.purgar_pendientes_vencidos()
            except Exception as e:
                logger.warning("Error en purgar_pendientes_vencidos en crear_reserva_estudiante: %s", e)

            id_horario = request.data.get('id_horario')
            id_activo = request.data.get('id_activo')
            acepta_dj = request.data.get('acepto_declaracion_jurada', False)

            if not all([id_horario, id_activo]):
                return Response({'error': "Faltan campos: id_horario, id_activo."}, status=400)

            if not acepta_dj:
                return Response({'error': "Debe aceptar la declaración jurada para reservar."}, status=400)

            # PBI-07: Verificar penalización activa antes de crear la reserva
            ahora = timezone.now()
            penalizacion_activa = Penalizacion.objects.filter(
                id_usuario=usuario,
                fecha_fin__gt=ahora
            ).order_by('-fecha_fin').first()
            if penalizacion_activa:
                dias_restantes = (penalizacion_activa.fecha_fin.date() - ahora.date()).days
                return Response({
                    'codigo': 'PENALIZACION_ACTIVA',
                    'error': (
                        f"Tienes una penalización activa hasta el "
                        f"{penalizacion_activa.fecha_fin.strftime('%d/%m/%Y')} "
                        f"({dias_restantes} día(s) restante(s)). "
                        f"No puedes realizar reservas."
                    ),
                    'fecha_fin_penalizacion': penalizacion_activa.fecha_fin.isoformat(),
                    'motivo': penalizacion_activa.motivo or 'No-show',
                }, status=403)

            try:
                reserva, error = reserva_service.crear_reserva_estudiante(
                    usuario=usuario,
                    id_horario=id_horario,
                    id_activo=id_activo,
                    acepta_dj=acepta_dj
                )
            except Exception as e:
                # S-4 (CWE-209): registrar el detalle internamente, no exponerlo.
                logger.error("Error al crear reserva de estudiante: %s", e, exc_info=True)
                return Response({'error': "Error interno del servidor. Intente más tarde."}, status=500)

            if error:
                return Response({'error': error}, status=409)

            try:
                lab_id = HorarioDisponible.objects.values_list('id_laboratorio_id', flat=True).get(pk=id_horario)
                cache.delete(f"activos_lab_{lab_id}_{id_horario}")
            except Exception:
                pass

            return Response({
                'mensaje': 'Reserva de equipo creada con éxito.',
                'reserva_id': reserva.id_reserva
            }, status=201)
        else:
            return Response({'error': f"El rol '{rol}' no está autorizado para crear reservas."}, status=403)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def cancelar_reserva(request, id_usuario, id_reserva):
    """Cancela una reserva existente a través de la ruta anidada del usuario."""
    # Validar permisos: el usuario autenticado debe coincidir con el id_usuario de la ruta (o ser admin/jefatura)
    if str(request.user.id_usuario) != str(id_usuario) and request.user.id_rol.nombre not in ('admin_lab', 'jefatura'):
        return Response({'error': 'No tienes permiso para cancelar esta reserva.'}, status=403)

    try:
        reserva = Reserva.objects.get(pk=id_reserva)
    except Reserva.DoesNotExist:
        return Response({'error': "Reserva no encontrada."}, status=404)

    # Validar que la reserva en efecto pertenezca al usuario del path
    if str(reserva.id_usuario_id) != str(id_usuario):
        return Response({'error': "La reserva especificada no pertenece al usuario indicado."}, status=400)

    if reserva.estado not in ('Programada', 'Pendiente'):
        return Response({'error': f"Solo puedes cancelar reservas en estado Programada o Pendiente. Estado actual: {reserva.estado}"}, status=400)

    motivo = request.data.get('motivo', '').strip()
    observacion = motivo if motivo else "Reserva cancelada por el usuario."
    es_cancelacion_por_admin = str(request.user.id_usuario) != str(id_usuario)

    lab_id = None
    try:
        with transaction.atomic():
            estado_anterior = reserva.estado
            reserva.estado = 'Cancelada'
            reserva.updated_at = timezone.now()  # Forzar actualización de timestamp
            reserva.save()

            # ID-03 / PBI-20: Liberar el horario usando la fuente de verdad
            # Bloqueamos el registro del HorarioDisponible para evitar race conditions al recalcular
            horario_id = reserva.id_horario_id
            if horario_id:
                horario_locked = HorarioDisponible.objects.select_for_update().get(pk=horario_id)
                lab_id = horario_locked.id_laboratorio_id
                from ..services.reserva_service import recalcular_capacidad
                recalcular_capacidad(horario_locked)

            # Log de la cancelación usando ID directo
            try:
                HistorialReserva.objects.create(
                    reserva=reserva,
                    estado_anterior=estado_anterior,
                    estado_nuevo='Cancelada',
                    usuario_cambio_id=request.user.id_usuario,
                    observacion=observacion,
                )
            except Exception as log_err:
                logger.warning("No se pudo crear el historial de reserva: %s", log_err)

        # Invalidar caché de activos tras commit exitoso
        if reserva.id_horario_id and lab_id:
            cache.delete(f"activos_lab_{lab_id}_{reserva.id_horario_id}")

        # Notificar al titular si el admin cancela una reserva ajena
        if es_cancelacion_por_admin:
            try:
                from ..services.reserva_service import enviar_notificacion_email
                titular = reserva.id_usuario
                horario = reserva.id_horario
                laboratorio_nombre = horario.id_laboratorio.nombre if horario else 'el laboratorio'
                fecha = horario.fecha.strftime('%d/%m/%Y') if horario else '—'
                hora_inicio = horario.hora_inicio.strftime('%H:%M') if horario else '—'
                hora_fin = horario.hora_fin.strftime('%H:%M') if horario else '—'
                asunto = "Tu reserva en LabSync ha sido cancelada"
                cuerpo = (
                    f"Hola {titular.nombre},\n\n"
                    f"Tu reserva en {laboratorio_nombre} programada para el {fecha} "
                    f"de {hora_inicio} a {hora_fin} ha sido cancelada por el administrador.\n\n"
                    f"Motivo: {motivo if motivo else 'No especificado'}\n\n"
                    f"Si tienes alguna consulta, contáctate con el área de laboratorios UNTELS."
                )
                enviar_notificacion_email(titular.email, asunto, cuerpo)
            except Exception as email_err:
                logger.warning("No se pudo enviar email de cancelación al titular: %s", email_err)

        return Response({'mensaje': 'Reserva cancelada correctamente.'})
    except Exception as e:
        # S-4 (CWE-209): no exponer detalles internos de la excepción al cliente.
        logger.error("Error al cancelar reserva %s: %s", id_reserva, e, exc_info=True)
        return Response({
            'error': "Error interno al procesar la cancelación."
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminOrJefatura])  # ID-05: solo admin/jefatura
def marcar_asistencia(request, id_reserva):
    """PBI-06: Administrador marca asistencia o No-Show."""
    # L-4 (CWE-704): coerción explícita — un string como "false" es truthy en Python
    # y marcaría Completada por error.
    raw_asistio = request.data.get('asistio', False)
    if isinstance(raw_asistio, str):
        asistio = raw_asistio.strip().lower() in ('true', '1', 'yes', 'si', 'sí')
    else:
        asistio = bool(raw_asistio)
    ok, error = reserva_service.marcar_asistencia(id_reserva, asistio)
    
    if not ok:
        return Response({'error': error}, status=400)
        
    estado = 'Completada' if asistio else 'No-show'
    return Response({'mensaje': f'Reserva marcada como {estado}.'})


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminOrJefatura])  # ID-06: solo admin/jefatura
def get_historial_reservas(request):
    """Devuelve el historial global de cambios de estado de todas las reservas."""
    from ..models import HistorialReserva
    from ..serializers.reserva_serializers import HistorialReservaSerializer
    
    # A-2: select_related evita N+1 al serializar cada entrada del historial
    historial = (
        HistorialReserva.objects
        .all()
        .select_related(
            'reserva',
            'reserva__id_usuario',
            'reserva__id_horario',
            'reserva__id_horario__id_laboratorio',
        )
        .order_by('-fecha_cambio')[:100]
    )
    serializer = HistorialReservaSerializer(historial, many=True)
    return Response(serializer.data)


# Q-3: cabeceras y construcción de filas compartidas por los exportadores CSV/XLSX.
_CABECERAS_REPORTE = [
    'Titular', 'Rol', 'Laboratorio', 'Fecha', 'Hora Inicio', 'Hora Fin',
    'Estado Reserva', 'Asistencia', 'Cantidad Alumnos', 'Motivo Cancelación',
]


def _filtrar_reservas_reporte(request):
    """Q-3: queryset de reservas con los filtros opcionales de fecha/laboratorio,
    común a los exportadores de Jefatura."""
    reservas = Reserva.objects.select_related(
        'id_usuario', 'id_usuario__id_rol', 'id_horario', 'id_horario__id_laboratorio', 'asistencia'
    ).order_by('-created_at')

    fecha_inicio = request.query_params.get('fecha_inicio')
    fecha_fin = request.query_params.get('fecha_fin')
    laboratorio_id = request.query_params.get('laboratorio_id')

    if fecha_inicio:
        reservas = reservas.filter(id_horario__fecha__gte=fecha_inicio)
    if fecha_fin:
        reservas = reservas.filter(id_horario__fecha__lte=fecha_fin)
    if laboratorio_id:
        reservas = reservas.filter(id_horario__id_laboratorio_id=laboratorio_id)
    return reservas


def _fila_reporte(r):
    """Q-3: convierte una Reserva en la fila de 10 columnas del reporte."""
    titular = r.id_usuario.nombre if r.id_usuario else 'N/A'
    rol = r.id_usuario.id_rol.nombre if r.id_usuario and r.id_usuario.id_rol else 'N/A'

    horario = r.id_horario
    if horario:
        laboratorio = horario.id_laboratorio.nombre if horario.id_laboratorio else 'N/A'
        fecha = horario.fecha.strftime('%Y-%m-%d')
        hora_inicio = horario.hora_inicio.strftime('%H:%M')
        hora_fin = horario.hora_fin.strftime('%H:%M')
    else:
        laboratorio, fecha, hora_inicio, hora_fin = 'N/A', 'N/A', 'N/A', 'N/A'

    estado = r.estado

    if hasattr(r, 'asistencia') and r.asistencia:
        asistencia = 'Sí' if r.asistencia.asistio else 'No'
    elif estado == 'Completada':
        asistencia = 'Sí'
    elif estado == 'No-show':
        asistencia = 'No'
    else:
        asistencia = 'N/A'

    motivo_cancel = 'N/A'
    if estado == 'Cancelada':
        historial = r.historialreserva_set.filter(estado_nuevo='Cancelada').first()
        if historial:
            motivo_cancel = historial.observacion or 'N/A'

    return [
        titular, rol, laboratorio, fecha, hora_inicio, hora_fin,
        estado, asistencia, r.cantidad_alumnos, motivo_cancel,
    ]


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsJefatura])
def exportar_csv_jefatura(request):
    """PBI-10: Exportar listado de reservas y asistencias a CSV para Jefatura."""
    reservas = _filtrar_reservas_reporte(request)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="reporte_reservas.csv"'

    # Añadir BOM para que Excel detecte UTF-8
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(_CABECERAS_REPORTE)
    for r in reservas:
        writer.writerow(_fila_reporte(r))

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsJefatura])
def exportar_xlsx_jefatura(request):
    """PBI-10: Exportar listado de reservas y asistencias a XLSX para Jefatura con estilos."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    import io

    reservas = _filtrar_reservas_reporte(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Reservas"

    ws.append(_CABECERAS_REPORTE)

    # Estilos para la cabecera (Azul oscuro y texto blanco)
    fill_color = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_blanca = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill_color
        cell.font = font_blanca

    for r in reservas:
        ws.append(_fila_reporte(r))

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        column = col[0].column_letter
        max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        ws.column_dimensions[column].width = max_length + 2

    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer, 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_reservas.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsJefatura])
def metricas_jefatura(request):
    """PBI-17: Dashboard de métricas operativas para Jefatura."""
    
    # Reservas por laboratorio
    reservas_por_lab = (
        Reserva.objects.filter(id_horario__isnull=False)
        .values('id_horario__id_laboratorio__nombre')
        .annotate(total=Count('id_reserva'))
        .order_by('-total')
    )
    
    # Formatear
    reservas_laboratorio = [
        {'laboratorio': item['id_horario__id_laboratorio__nombre'], 'total_reservas': item['total']}
        for item in reservas_por_lab
    ]

    # Tasa de cancelaciones
    total_reservas = Reserva.objects.count()
    canceladas = Reserva.objects.filter(estado='Cancelada').count()
    porcentaje_canceladas = round((canceladas / total_reservas * 100), 2) if total_reservas > 0 else 0
    tasa_cancelaciones = {
        'canceladas': canceladas,
        'total': total_reservas,
        'porcentaje': porcentaje_canceladas
    }

    # No-shows por usuario (top 10)
    noshows_por_usuario = (
        Reserva.objects.filter(estado='No-show')
        .values('id_usuario__nombre')
        .annotate(noshow_count=Count('id_reserva'))
        .order_by('-noshow_count')[:10]
    )
    noshow_usuarios = [
        {'nombre_usuario': item['id_usuario__nombre'], 'noshow_count': item['noshow_count']}
        for item in noshows_por_usuario
    ]

    # Distribución de estados
    estados = Reserva.objects.values('estado').annotate(total=Count('id_reserva'))
    estado_distribucion = {item['estado']: item['total'] for item in estados}

    return Response({
        'reservas_por_laboratorio': reservas_laboratorio,
        'tasa_cancelaciones': tasa_cancelaciones,
        'noshow_por_usuario': noshow_usuarios,
        'estado_distribucion': estado_distribucion
    })
